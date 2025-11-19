import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score
import openpyxl
from sklearn import linear_model
from sklearn.ensemble import RandomForestRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.neural_network import MLPRegressor
from itertools import combinations


def mape(true, pred):
    true = np.array(true).reshape(-1, 1)
    pred = np.array(pred).reshape(-1, 1)
    diff = np.abs(true - pred)
    return np.mean(diff / np.abs(true))


def traversing_the_basic_quantity(read_path, write_path, target_label, get_all_combination_label, model,
                                  data_to_drop=['E_C', 'F', 'T'], base_combination=[0, 3, 5], test_size=0.25,
                                  random_state_value=0, data_set=1):
    """
    :param read_path: 需要读取的文件存放路径
    :param write_path: 需要写入的文件存放路径
    :param target_label: 可选0/1/2，分别代表：0预测能耗 1预测推力 2预测扭矩
    :param get_all_combination_label: 可选0/1，分别代表：0 输入指定基本量组合；1 遍历所有基本量组合
    :param model: 希望用于检验pi量组合的sklearn模型
    :param data_to_drop: 三个潜在目标量的命名，默认按照能耗、推力、扭矩的顺序
    :param base_combination: get_all_combination_label=0时使用，代表选择的基本量组合
    :param test_size: 检验模型时训练集占比
    :param random_state_value: 随机因子
    :param data_set: 所用数据集，可选0/1/2，分别代表：0大伙房 1天津9 2天津3
    :return: 只有get_all_combination_label=0时有返回值，返回对应基本量下的pi量和相关目标量、还原目标量会用到的系数
    """
    # sheet_list = ['大伙房原始', '天津9原始', '天津3原始', '大伙房量纲', '天津量纲']
    sheet_list = ['大伙房原始', '北京9原始', '天津3原始', '大伙房量纲', '天津量纲']
    data_read = pd.read_excel(read_path, sheet_name=sheet_list[data_set])
    del data_to_drop[target_label]
    data_use = data_read.drop(columns=data_to_drop)
    num_of_feature = np.shape(data_use)[1] - 1  # 定位目标量量纲的位置

    matrix_to_drop = [num_of_feature, num_of_feature + 1, num_of_feature + 2]
    del matrix_to_drop[target_label]
    if data_set == 0:
        matrix_read = pd.read_excel(read_path, sheet_name=sheet_list[3], header=None)  # 读取特征的量纲信息
    else:
        matrix_read = pd.read_excel(read_path, sheet_name=sheet_list[4], header=None)  # 读取特征的量纲信息
    matrix_read = matrix_read.drop(columns=matrix_to_drop)

    pass_col = []  # 把全是0的列(本身就是无量纲的)的列号记录下来
    for col_num in range(np.shape(matrix_read)[1]):
        if all(item == 0 for item in list(matrix_read.iloc[:, col_num])):
            pass_col.append(col_num)
    pass_col.append(col_num)  # 除去目标量的量纲
    all_col = [i for i in range(np.shape(matrix_read)[1])]  # 所有特征的列号
    use_col = [i for num, i in enumerate(all_col) if num not in pass_col]  # 删除所有列号中无量纲的列号值

    if get_all_combination_label == 1:
        possible_combination = []  # 存储可能的基本量的选取方式，存储的基本量是其对应的原始数据中的列号
        for first_base in range(len(use_col) - 2):
            for sec_base in range(first_base + 1, len(use_col) - 1):
                for third_base in range(sec_base + 1, len(use_col)):
                    choose_matrix = np.matrix(
                        matrix_read.iloc[:, [use_col[first_base], use_col[sec_base], use_col[third_base]]])
                    if np.linalg.det(choose_matrix) == 0:
                        continue
                    possible_combination.append([use_col[first_base], use_col[sec_base], use_col[third_base]])
    else:
        possible_combination = [base_combination]

    possible_coefficient = []  # 存储所有可能基本量选取方式下，其他量无量纲化时所需要的基本量的系数(系数取相反数才是应该乘的系数)
    for combination in possible_combination:
        not_base_col = [i for num, i in enumerate(use_col) if i not in combination]
        not_base_col.append(col_num)  # 这里需要加入预测的目标量
        base_matrix = np.matrix(matrix_read.iloc[:, combination])
        single_solution_coefficients = []
        for single_col in not_base_col:
            single_dim = list(matrix_read.iloc[:, single_col])
            single_coefficient = list(np.linalg.solve(base_matrix, single_dim))
            single_solution_coefficients.append(single_coefficient)
        possible_coefficient.append(single_solution_coefficients)

    r2_pai_list = []  # 用于存储不同pai量方案下未还原的r2
    r2_real_list = []  # 用于存储不同pai量方案下还原后的r2
    mape_pai_list = []  # 用于存储不同pai量方案下未还原的mape
    mape_real_list = []  # 用于存储不同pai量方案下未还原的mape
    count = 0
    for coefficient, combination in zip(possible_coefficient, possible_combination):
        not_base_col = [i for num, i in enumerate(use_col) if i not in combination]
        not_base_col.append(col_num)
        base_col_1 = data_use.iloc[:, combination[0]]
        base_col_2 = data_use.iloc[:, combination[1]]
        base_col_3 = data_use.iloc[:, combination[2]]
        single_solution_pai = []  # 用于存储使用一套基本量无量纲化后的各特征形成的pai量数据
        for single_not_base_col, single_feature_coefficient in zip(not_base_col, coefficient):
            not_base_feature = data_use.iloc[:, single_not_base_col]
            single_pai = not_base_feature * (base_col_1 ** (-single_feature_coefficient[0])) * \
                         (base_col_2 ** (-single_feature_coefficient[1])) * (
                                     base_col_3 ** (-single_feature_coefficient[2]))
            single_pai = list(single_pai)
            single_solution_pai.append(single_pai)
        for i in range(len(pass_col) - 1):  # 将本身就是无量纲的量加入到无量纲化后的列表中
            single_solution_pai.insert(-1, list(data_use.iloc[:, pass_col[i]]))
        single_solution_pai.append(data_use.iloc[:, -1])  # 将真实目标值存入该列表，便于之后还原后对比
        for i in combination:
            single_solution_pai.append(list(data_use.iloc[:, i]))  # 将真实的基本量的值存入列表，便于之后还原
        single_solution_pai = np.array(single_solution_pai).transpose()

        # 使用得到的pai量训练模型
        X = single_solution_pai[:, : -5]  # 除开最后五列：目标pai值、目标真实值、三个基本量
        y = single_solution_pai[:, -5:]

        if get_all_combination_label == 0:
            return single_solution_pai, coefficient

        x_train, x_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state_value)
        clf = model
        clf.fit(x_train, y_train[:, 0])

        predict = clf.predict(x_test)
        trans_predict = predict * (y_test[:, 2] ** (coefficient[-1][0])) * \
                        (y_test[:, 3] ** (coefficient[-1][1])) * (y_test[:, 4] ** (coefficient[-1][2]))
        r2_pai = r2_score(y_test[:, 0], predict, sample_weight=None)
        r2_real = r2_score(y_test[:, 1], trans_predict, sample_weight=None)
        mape_pai = mape(y_test[:, 0], predict)
        mape_real = mape(y_test[:, 1], trans_predict)
        r2_pai_list.append(r2_pai)
        r2_real_list.append(r2_real)
        mape_pai_list.append(mape_pai)
        mape_real_list.append(mape_real)
        count = count + 1
        print(
            '正在计算基本量组合：' + str(combination) + ' 下的结果，总体进度为：' + str(int(count * 100 / len(possible_combination))) + '%')

    if get_all_combination_label == 1:
        title_list = ['刀盘直径', '顶部埋深变正m', '掘进速度mm/min', '逐环平均转速rpm', '土仓压力左右平均值MPa', '弹性模量 Mpa', '泊松比', '容重',
                      '粘聚力 kPa', '静止土压力系数', '内摩擦角 °', '岩土承载力 kPa', '原始r2', '还原后r2', '原始mape', '还原后mape']

        wb_w = openpyxl.load_workbook(write_path, data_only=True)
        if target_label == 1:
            wsheet = wb_w.create_sheet('预测推力_lr')
        elif target_label == 2:
            wsheet = wb_w.create_sheet('预测扭矩_lr')
        else:
            wsheet = wb_w.create_sheet('预测能耗_lr')

        for col in range(len(title_list)):  # 写入表头
            wsheet.cell(1, col + 1, title_list[col])

        for row in range(len(possible_combination)):
            for col in range(12):
                if col in possible_combination[row]:
                    wsheet.cell(row + 2, col + 1, 1)
                else:
                    wsheet.cell(row + 2, col + 1, 0)
            wsheet.cell(row + 2, 13, r2_pai_list[row])
            wsheet.cell(row + 2, 14, r2_real_list[row])
            wsheet.cell(row + 2, 15, mape_pai_list[row])
            wsheet.cell(row + 2, 16, mape_real_list[row])

        wb_w.save(write_path)


def pick_big_pi_to_test(pi_combinations_dataframe, y, write_path_big_pi, coefficient, model, big_pi_choose_num=6,
                        combination_nums=8400000):
    """
    :param pi_combinations_dataframe: 构造好的作为候选的大pi量
    :param y: 预测目标量，共5列，格式是traversing_the_basic_quantity函数第一个返回值的最后五列
    :param write_path_big_pi: 记录大pi量计算结果的路径
    :param coefficient: 还原目标量会用到的系数
    :param model: 选用的评价模型
    :param big_pi_choose_num: 单个大pi量组合需要从候选集中选择的大pi量个数
    :param combination_nums:预估的组合数，用于输出计算进度
    :return: 无返回值
    """
    # 挑选上面的大pi组合中的n=6项用于训练预测模型
    big_pi_combination_to_write = []
    big_pi_combination_to_write_distributed = []

    num_of_candidate = np.shape(pi_combinations_dataframe)[1]

    index_list = list(range(0, num_of_candidate))
    # 通过迭代器的方式可以避免直接存储上亿的组合，然后内存崩溃
    index_choose_from_combination_tuple = combinations(index_list, big_pi_choose_num)

    score_name_list = ['r2_pai', 'r2_real', 'mape_pai', 'mape_real']
    write_columns = list(range(0, big_pi_choose_num + 4))
    for i in range(big_pi_choose_num + 4):
        if i < big_pi_choose_num:
            write_columns[i] = 'big_pi' + str(i + 1)
        else:
            write_columns[i] = score_name_list[i - big_pi_choose_num]
    count = 1

    while True:
        try:
            single_index = next(index_choose_from_combination_tuple)
            single_big_pi_combination = pi_combinations_dataframe.iloc[:, list(single_index)]
            single_big_pi_combination_to_write = list(single_big_pi_combination.columns)

            r2_pai, r2_real, mape_pai, mape_real = train_and_test_pi(single_big_pi_combination, y, coefficient, model)

            single_big_pi_combination_to_write.append(r2_pai)
            single_big_pi_combination_to_write.append(r2_real)
            single_big_pi_combination_to_write.append(mape_pai)
            single_big_pi_combination_to_write.append(mape_real)
            big_pi_combination_to_write.append(single_big_pi_combination_to_write)
            big_pi_combination_to_write_distributed.append(single_big_pi_combination_to_write)  # 用于分块存储

            if count % 1000 == 0:
                print('计算进度：%.3f %%' % (count / combination_nums * 100))
            if count % 1000000 == 0:
                distributed_to_wirte = pd.DataFrame(big_pi_combination_to_write_distributed, columns=write_columns)
                distributed_to_wirte.to_csv(write_path_big_pi + '\大pi方案_除法' + str(round(count / 500000)) + '.csv')
                big_pi_combination_to_write_distributed = []

            count += 1
        except StopIteration:
            break

    distributed_to_wirte = pd.DataFrame(big_pi_combination_to_write_distributed, columns=write_columns)
    distributed_to_wirte.to_csv(write_path_big_pi + '\大pi方案_除法' + str(int(count % 500000)) + '.csv')

    final_to_wirte = pd.DataFrame(big_pi_combination_to_write, columns=write_columns)
    final_to_wirte.to_csv(write_path_big_pi + '\大pi方案_除法.csv')


def compute_big_pi(data_use, expression_list):
    """
    :param data_use: 用于构建大pi的小pi数据
    :param expression_list: 构建大pi的方案，用字符串表示，对一个大pi，表达式中用p代表pi，'p1*p2'代表由pi1和pi2相乘得到的大pi
    :return: 构建好的大pi数据
    """
    import math
    big_pi = []
    data = data_use  # data是在后面解析表达式中使用，这里编译器会误认为没有使用
    for expression_read in expression_list:
        location_mark = 0
        p_mark = 0
        location = ''
        expression = ''
        for item in expression_read:  # 单个大pi计算
            if item == 'p' or item == 'P':  # 遇到p代表大pi，p本身不写入真正的表达式
                location_mark = 1
                p_mark = 1
                continue
            if not item.isdigit() and p_mark == 1:  # 防止把其他p相关字符作为pi量
                expression += 'p'
                location_mark = 0
            elif item.isdigit() and p_mark == 1:  # 上一轮就是p代表pi量的情况
                expression += 'data[:, '

            if location_mark == 1 and item.isdigit():
                location += item
            elif location_mark == 1:
                expression += location
                expression += ']'
                expression += item
                location = ''
                location_mark = 0
            else:
                expression += item

            if not (item == 'p' or item == 'P'):  # 用于下一轮判断上一轮是否是p
                p_mark = 0

        if location_mark == 1:  # 最后一位是数字的情况
            expression += location
            expression += ']'
        single_big_pi = eval(expression)
        big_pi.append(single_big_pi)
    big_pi = np.array(big_pi).transpose()
    return big_pi


def train_and_test_pi(data_x, data_y, coefficient, model, test_size=0.25, train_size=0.75, random_state_value=0):
    """
    :param data_x: 大pi或小pi的输入数据
    :param data_y: 预测目标量，共5列，格式是traversing_the_basic_quantity函数第一个返回值的最后五列
    :param coefficient: 还原目标量会用到的系数(traversing_the_basic_quantity计算得到)
    :param model: 选用的评价模型
    :param test_size: 用于评价模型的测试集占比或个数
    :param train_size: 训练集占比或个数
    :param random_state_value:随机种子
    :return:还原前后的r2和还原前后的mape
    """
    x_train, x_test, y_train, y_test = train_test_split(data_x, data_y, test_size=test_size, train_size=train_size,
                                                        random_state=random_state_value)
    clf = model
    clf.fit(x_train, y_train[:, 0])
    predict = clf.predict(x_test)
    trans_predict = predict * (y_test[:, 2] ** (coefficient[-1][0])) * \
                    (y_test[:, 3] ** (coefficient[-1][1])) * (y_test[:, 4] ** (coefficient[-1][2]))
    # 通常需要加上 * x_test[:, 0] $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

    r2_pai = round(r2_score(y_test[:, 0], predict, sample_weight=None), 4)
    r2_real = round(r2_score(y_test[:, 1], trans_predict, sample_weight=None), 4)
    mape_pai = mape(y_test[:, 0], predict)
    mape_real = mape(y_test[:, 1], trans_predict)
    return r2_pai, r2_real, mape_pai, mape_real


def train_and_test(read_path, sheet_name, target_label, model, test_size=0.25, train_size=0.75,
                   random_state_value=0, shuffle=True):
    """
    :param read_path: 数据存放路径(包含到.xlsx)
    :param sheet_name: 读取工作表的名称
    :param target_label: 可选0/1/2，分别代表：0预测能耗 1预测推力 2预测扭矩
    :param model: 拟合模型的选择
    :param test_size: 测试集占比或个数
    :param train_size: 训练集占比或个数
    :param random_state_value: 随机种子
    :param shuffle: 是否打乱数据集
    :return: 测试集上模型的r2和mape
    """
    data = pd.read_excel(read_path, sheet_name=sheet_name)
    data_x = data.iloc[:, :-3]
    data_y = data.iloc[:, target_label - 3]
    x_train, x_test, y_train, y_test = train_test_split(data_x, data_y, test_size=test_size, train_size=train_size,
                                                        random_state=random_state_value, shuffle=shuffle)
    clf = model
    clf.fit(x_train, y_train)
    predict = clf.predict(x_test)
    r2 = round(r2_score(y_test, predict, sample_weight=None), 4)
    mape_out = mape(y_test, predict)
    return r2, mape_out


# ————————————————————————————————————————————————路径设置与模型设置
# read_path_ = r"D:\p-excel\研二下工作\师姐工作改进\使用数据\常用数据.xlsx"
# write_path_ = r'D:\p-excel\研二下工作\师姐工作改进\output\pi方案.xlsx'
#
rfr = RandomForestRegressor(max_depth=10, random_state=0, criterion='mae')
lr = LinearRegression(fit_intercept=True)  # 建立用于检验pai量方案效果的模型
knn = KNeighborsRegressor(weights="uniform")
lasso1 = linear_model.Lasso(alpha=1e-10, fit_intercept=False, positive=True,
                            random_state=0, max_iter=500000, selection='random', tol=1e-20)
lasso2 = linear_model.Lasso(alpha=1e-10, fit_intercept=True, positive=False,
                            random_state=0, max_iter=500000, selection='random', tol=1e-20)
ann = MLPRegressor(hidden_layer_sizes=(20, 10), activation='relu', solver='adam', max_iter=5000)

# ————————————————————————————————————————————————调用获取小pi的函数
# read_path_ = r"D:\p-excel\博一上工作\计算内容\使用数据\校正-天津9.xlsx"
# read_path_ = r'D:\p-excel\博一上工作\小论文相关\小周论文整理\待发表2——载荷\使用数据\dataset_all.xlsx'
write_path_ = r'C:\Users\lsl\Desktop\tool.xlsx'
read_path_ = r'D:\p-excel\博一下工作\计算内容\数据\常用数据.xlsx'

pi_data, coefficient_ = traversing_the_basic_quantity(read_path_, write_path_, target_label=1,
                                                      get_all_combination_label=0, model=lr, data_set=1, base_combination=[0, 3, 5])
pd.DataFrame(pi_data).to_csv(r'C:\Users\lsl\Desktop\tool.csv')
# 通常需要改target_label、data_set、base_combination$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

# pi_without_target = pi_data[:, : -5]  # 只包含作为输入的pi量
# pi_num = np.shape(pi_without_target)[1]  # 作为输入的小pi数量
# y_ = pi_data[:, -5:]  # 作为核验大pi组合的目标量
# # y_[:, 0] = y_[:, 0] / pi_without_target[:, 0]  # 作为核验大pi组合的目标量/大伙房使用
# 通常需要加上面这一行$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

# ————————————————————————————————————————————————遍历规则下的各种大pi组合
# pi_combinations = []  # 备选组合数据集
# pi_combination_names = []
#
# for first_item in range(pi_num):  # 添加由2个pi量组成大pi的情况
#     for sec_item in range(first_item + 1, pi_num):
#         # pi_combinations.append(pi_without_target[:, first_item] * pi_without_target[:, sec_item])
#         # pi_combination_names.append('p' + str(first_item) + '*' + 'p' + str(sec_item))
#         pi_combinations.append(pi_without_target[:, first_item] / pi_without_target[:, sec_item])  # 考虑有除法的情况
#         pi_combination_names.append('p' + str(first_item) + '/' + 'p' + str(sec_item))
#
# for item in range(pi_num):  # 添加由1个pi量组成大pi的情况
#     pi_combinations.append(pi_without_target[:, item])
#     pi_combination_names.append('p' + str(item))
#
# pi_combinations_dataframe_ = pd.DataFrame(np.transpose(pi_combinations), columns=pi_combination_names)
# write_path_big_pi_ = r'D:\p-excel\研二下工作\师姐工作改进\output'
# pick_big_pi_to_test(pi_combinations_dataframe_, y_, write_path_big_pi_, coefficient_, lr)  # 调用计算大pi的函数


# ————————————————————————————————————————————————根据设定组合方式获取大pi
# expression_list_ = ['p2', '(1+p7)*p3*p0', 'p1*np.tan(p8/180 * math.pi)', 'p5', 'p1*p4', 'p1/(2*(1+p6))']  # 师姐大pi
# expression_list_ = ['p0*p2', 'p0*p3', 'p0*p6', 'p3*p8', 'p6*p8', 'p7']  # 乘法组合准确度最高的
# big_pi_ = compute_big_pi(pi_without_target, expression_list_)


# ————————————————————————————————————————————————计算pi相关数据的表现
# r2_pai_, r2_real_, mape_pai_, mape_real_ = train_and_test_pi(big_pi_, y_, coefficient_, lr)


# ————————————————————————————————————————————————计算原始数据的表现
# r2_, mape_ = train_and_test(read_path_, '天津9原始', 1, lr, shuffle=True)


# ————————————————————————————————————————————————不同数量训练样本下模型准确度
# model_list = [lasso1, lr, knn, rfr, ann]
# r2_list = []
# # num_list = [5, 10, 20, 30, 50, 100, 150, 200, 300, 500]
# # num_list = [5, 10, 20, 30, 50, 100, 150, 200, 500, 1000]
# num_list = [5, 10, 15, 20, 30, 50, 100]
# for num in num_list:
#     r2_row = []
#     for model in model_list:
#         # r2_, mape_ = train_and_test(read_path_, '天津3原始', 1, model, train_size=num, test_size=80)
#         # 通常需要改sheet的名称、目标量、test_size $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#         r2_pai_, r2_real_, mape_pai_, mape_real_ = train_and_test_pi(big_pi_, y_, coefficient_, model, train_size=num,
#                                                                       test_size=80)
#         # 通常需要改pi或大pi、test_size $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#         r2_row.append(r2_real_)
#         # 通常需要改r2的内容$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#     r2_list.append(r2_row)
#
# pd.DataFrame(r2_list).to_csv(r'D:\p-excel\研二下工作\师姐工作改进\output\训练样本数量讨论\TJ3_大pi_训练数量讨论.csv')
# # 通常需要改存储文件名称$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$


# ——————————————————————————————————————————————不同数据集下模型的准确度
# read_path_ = r'D:\p-excel\研二下工作\师姐工作改进\使用数据\周皓师兄数据-逐环平均.xlsx'
# r2_list = []
# for model in model_list:
#     r2_, mape_ = train_and_test(read_path_, 'try', 0, model, shuffle=False, random_state_value=0)
#     r2_list.append(r2_)
# pd.DataFrame(r2_list).to_csv(r'D:\p-excel\研二下工作\师姐工作改进\output\训练样本数量讨论\tool.csv')
# print(r2_list)


