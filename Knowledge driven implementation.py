import numpy as np
import pandas as pd
import openpyxl
from itertools import combinations
from sklearn.linear_model import LinearRegression


def traversing_the_basic_quantity(read_path, write_path, target_label, get_all_combination_label, model,
                                  data_to_drop=['E_C', 'F', 'T'], base_combination=[0, 3, 5], test_size=0.25,
                                  random_state_value=0, data_set=1):
    """
    :param read_path: The file storage path that needs to be read
    :param write_path: The file storage path that needs to be written
    :param target_label: Optional 0/1/2, representing: 0 predicted energy consumption, 1 predicted thrust, 2 predicted torque
    :param get_all_combination_label: Optional 0/1, representing: 0 input specified basic quantity combination; Traverse all combinations of fundamental quantities
    :param model: Hope to use sklearn model for testing pi quantity combinations
    :param data_to_drop: The naming of the three potential target quantities is based on the order of energy consumption, thrust, and torque by default
    :param base_combination: Used when get-all_combination_1abel=0, representing the selected combination of fundamental quantities
    :param test_size: Proportion of training set when testing the model
    :param random_state_value: random factor
    :param data_set: The dataset used can be selected as 0/1/2, representing: 0 Dahuofang 1 Tianjin 9 2 Tianjin 3
    :return: Only when get-all_combination_1abel=0, there is a return value, which returns the pi value under the corresponding basic quantity, the related target quantity, and the coefficients used to restore the target quantity
    """
    # sheet_list = ['DHF', 'TJ9', 'TJ3', 'DHF_Dim', 'TJ_Dim']
    sheet_list = ['DHF', 'TJ9', 'TJ3', 'DHF_Dim', 'TJ_Dim']
    data_read = pd.read_excel(read_path, sheet_name=sheet_list[data_set])
    del data_to_drop[target_label]
    data_use = data_read.drop(columns=data_to_drop)
    num_of_feature = np.shape(data_use)[1] - 1  # Locate the position of the target dimension

    matrix_to_drop = [num_of_feature, num_of_feature + 1, num_of_feature + 2]
    del matrix_to_drop[target_label]
    if data_set == 0:
        matrix_read = pd.read_excel(read_path, sheet_name=sheet_list[3], header=None)  # Read dimensional information of features
    else:
        matrix_read = pd.read_excel(read_path, sheet_name=sheet_list[4], header=None)  # Read dimensional information of features
    matrix_read = matrix_read.drop(columns=matrix_to_drop)

    pass_col = []  # Record the column numbers of columns that are all zeros (already dimensionless)
    for col_num in range(np.shape(matrix_read)[1]):
        if all(item == 0 for item in list(matrix_read.iloc[:, col_num])):
            pass_col.append(col_num)
    pass_col.append(col_num)  # Excluding the dimension of the target quantity
    all_col = [i for i in range(np.shape(matrix_read)[1])]  # Column numbers for all features
    use_col = [i for num, i in enumerate(all_col) if num not in pass_col]  # Delete all dimensionless column number values from all column numbers

    if get_all_combination_label == 1:
        possible_combination = []  # The selection method for storing possible basic quantities, where the stored basic quantity is the column number in the corresponding raw data
        for first_base in range(len(use_col) - 2):
            for sec_base in range(first_base + 1, len(use_col) - 1):
                for third_base in range(sec_base + 1, len(use_col)):
                    choose_matrix = np.matrix(
                        matrix_read.iloc[:, [use_col[first_base], use_col[sec_base], use_col[third_base]]])
                    if np.linalg.det(choose_matrix) == 0:
                        continue
                    possible_combination.append([use_col[first_base], use_col[sec_base], use_col[third_base]])
    else:
        possible_combination = [base_combination]

    possible_coefficient = []  # Store the coefficients of the fundamental quantities required for dimensionless transformation of other quantities under all possible fundamental quantity selection methods (the coefficients should be multiplied by the opposite number)
    for combination in possible_combination:
        not_base_col = [i for num, i in enumerate(use_col) if i not in combination]
        not_base_col.append(col_num)  # Need to add the predicted target quantity here
        base_matrix = np.matrix(matrix_read.iloc[:, combination])
        single_solution_coefficients = []
        for single_col in not_base_col:
            single_dim = list(matrix_read.iloc[:, single_col])
            single_coefficient = list(np.linalg.solve(base_matrix, single_dim))
            single_solution_coefficients.append(single_coefficient)
        possible_coefficient.append(single_solution_coefficients)

    r2_pai_list = []  # Used to store unrestored r2 under different pair quantity schemes
    r2_real_list = []  # Used to store the restored r2 under different pair quantity schemes
    mape_pai_list = []  # Used to store unrestored mapes under different pai quantity schemes
    mape_real_list = []  # Used to store unrestored mapes under different pai quantity schemes
    count = 0
    for coefficient, combination in zip(possible_coefficient, possible_combination):
        not_base_col = [i for num, i in enumerate(use_col) if i not in combination]
        not_base_col.append(col_num)
        base_col_1 = data_use.iloc[:, combination[0]]
        base_col_2 = data_use.iloc[:, combination[1]]
        base_col_3 = data_use.iloc[:, combination[2]]
        single_solution_pai = []  
        for single_not_base_col, single_feature_coefficient in zip(not_base_col, coefficient):
            not_base_feature = data_use.iloc[:, single_not_base_col]
            single_pai = not_base_feature * (base_col_1 ** (-single_feature_coefficient[0])) * \
                         (base_col_2 ** (-single_feature_coefficient[1])) * (
                                     base_col_3 ** (-single_feature_coefficient[2]))
            single_pai = list(single_pai)
            single_solution_pai.append(single_pai)
        for i in range(len(pass_col) - 1):  
            single_solution_pai.insert(-1, list(data_use.iloc[:, pass_col[i]]))
        single_solution_pai.append(data_use.iloc[:, -1])  
        for i in combination:
            single_solution_pai.append(list(data_use.iloc[:, i])) 
        single_solution_pai = np.array(single_solution_pai).transpose()

        # Train the model using the obtained pai quantity
        X = single_solution_pai[:, : -5]  # Excluding the last five columns: target pai value, target true value, and three basic quantities
        y = single_solution_pai[:, -5:]

        if get_all_combination_label == 0:
            return single_solution_pai, coefficient

        x_train, x_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=random_state_value)
        clf = model
        clf.fit(x_train, y_train[:, 0])

        predict = clf.predict(x_test)
        trans_predict = predict * (y_test[:, 2] ** (coefficient[-1][0])) * \
                        (y_test[:, 3] ** (coefficient[-1][1])) * (y_test[:, 4] ** (coefficient[-1][2]))
        r2_pai = r2_score(y_test[:, 0], predict, sample_weight=None)
        r2_real = r2_score(y_test[:, 1], trans_predict, sample_weight=None)
        mape_pai = mape(y_test[:, 0], predict)
        mape_real = mape(y_test[:, 1], trans_predict)
        r2_pai_list.append(r2_pai)
        r2_real_list.append(r2_real)
        mape_pai_list.append(mape_pai)
        mape_real_list.append(mape_real)
        count = count + 1

    if get_all_combination_label == 1:
        title_list = ['刀盘直径', '顶部埋深变正m', '掘进速度mm/min', '逐环平均转速rpm', '土仓压力左右平均值MPa', '弹性模量 Mpa', '泊松比', '容重',
                      '粘聚力 kPa', '静止土压力系数', '内摩擦角 °', '岩土承载力 kPa', '原始r2', '还原后r2', '原始mape', '还原后mape']

        wb_w = openpyxl.load_workbook(write_path, data_only=True)
        if target_label == 1:
            wsheet = wb_w.create_sheet('预测推力_lr')
        elif target_label == 2:
            wsheet = wb_w.create_sheet('预测扭矩_lr')
        else:
            wsheet = wb_w.create_sheet('预测能耗_lr')

        for col in range(len(title_list)):  # 写入表头
            wsheet.cell(1, col + 1, title_list[col])

        for row in range(len(possible_combination)):
            for col in range(12):
                if col in possible_combination[row]:
                    wsheet.cell(row + 2, col + 1, 1)
                else:
                    wsheet.cell(row + 2, col + 1, 0)
            wsheet.cell(row + 2, 13, r2_pai_list[row])
            wsheet.cell(row + 2, 14, r2_real_list[row])
            wsheet.cell(row + 2, 15, mape_pai_list[row])
            wsheet.cell(row + 2, 16, mape_real_list[row])

        wb_w.save(write_path)

write_path_ = r'C:\Users\shangkai\tool.xlsx'
read_path_ = r'D:\p-excel\shangkai\tj9_use_for_pi.xlsx'
model = linear_model.Lasso(alpha=1e-10, fit_intercept=True, positive=False,
                            random_state=0, max_iter=500000, selection='random', tol=1e-20)
pi_data, coefficient_ = traversing_the_basic_quantity(read_path_, write_path_, target_label=1, model=model,
                                                      get_all_combination_label=0, data_set=1, base_combination=[0, 3, 5])
